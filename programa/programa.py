# Declaração de imports
import pandas as pd
from openpyxl import load_workbook
import os
from pathlib import Path

#Declaração de variáveis
info_dictionary = {}
path = Path(os.getcwd())
base_excel_path = str(path.parent).replace("\\", "/")
source_excel = 'ORIGEM.xlsx'
destiny_excel = 'DESTINO.xlsx'
source_excel_path = base_excel_path + '/excel/' + source_excel
destiny_excel_path =  base_excel_path + '/excel/' + destiny_excel
destiny_column_reference = "CNPJ"

#Processo responsável por pegar uma lista dos CNPJ's no arquivo excel de Origem.
source = pd.read_excel(source_excel_path, sheet_name = 0)
quantity_cnpj_source = source.CNPJ.count()

for line in range(quantity_cnpj_source):
    cnpj = source.loc[line, "CNPJ"]
    cnpj_replace = str(cnpj).replace(".", "").replace("/", "").replace("-", "")
    status = str(source.loc[line, "STATUS"])
    info_dictionary[cnpj_replace] = status

# print(info_dictionary)

# Processo responsável por realizar um de > para com o arquivo excel de Destino.
destiny = pd.read_excel(destiny_excel_path, sheet_name = 0)
quantity_customer_destiny = destiny.RESPONSÁVEL.count()

wbDestiny = load_workbook(destiny_excel_path)
wsDestiny = wbDestiny.active

for x in range(quantity_customer_destiny):
    if pd.isna(destiny.loc[x, destiny_column_reference]) != True:
        destiny_cnpj = str(int(destiny.loc[x, destiny_column_reference]))
        status_dict = info_dictionary.get(destiny_cnpj, None)

        if status_dict != None : 
            wsDestiny.cell(row=x+2, column=20, value=status_dict)

#Salvar o excel.
wbDestiny.save(base_excel_path + '/excel/' + destiny_excel)